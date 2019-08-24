import openpyxl
def get_excel(file_name,sheet):
        #打开Excel
        wb=openpyxl.load_workbook(file_name)
        #访问sheet页
        sheet=wb[sheet]
        #包含最小的行索引从1开始
        minRow = sheet.min_row
        print("最小行索引是：",minRow)
        maxRow = sheet.max_row
        print("最大行索引是",maxRow)
        minColumn=sheet.min_column
        print("最小列索引是",minColumn)
        max_Column = sheet.max_column
        print("最大列索引是：",max_Column)
get_excel('test_data.xlsx','test_data')

    # 循环提取Excel数据

