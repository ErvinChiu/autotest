from openpyxl import load_workbook
class DoExcel:
    def __init__(self,file_path,sheet_name):
        self.file_path = file_path
        self.sheet_name=sheet_name
    def get_excel(self):
        wb = load_workbook(self.file_path)
        sheet = wb[self.sheet_name]
        test_data = []
        for i in range(2,sheet.max_row+1):
            sub_data={}
            sub_data['id'] = sheet.cell(i,1).value
            sub_data['title'] = sheet.cell(i, 2).value
            sub_data['mothod'] = sheet.cell(i, 3).value
            sub_data['url'] = sheet.cell(i, 4).value
            sub_data['param'] = sheet.cell(i, 5).value
            test_data.append(sub_data)
        return test_data

if __name__=='__main__':
    test_data = DoExcel("test_data.xlsx","test_data").get_excel()
    for item in test_data:
        print(item)
